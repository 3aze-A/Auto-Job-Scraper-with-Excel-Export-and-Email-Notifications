from openpyxl import Workbook, load_workbook
from openpyxl import styles
from datetime import date
from openpyxl.styles import Border, Side
import os


def main(all_jobs):
    try:
        if not os.path.exists('job_listings.xlsx'):
            # create a new Excel sheet
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 100
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 30

            # Add 'JOB LISTINGS' title with formatting
            ws['A1'] = 'JOB LISTINGS'
            ws['A1'].font = styles.Font(name='Arial', bold=True, size=16, color='FFFFFF')
            ws['A1'].fill = styles.PatternFill('solid', start_color='2F5496')  # dark blue
            ws['A1'].alignment = styles.Alignment(horizontal='center')
            ws.merge_cells('A1:E1')

            last_line = 1
        else:
            wb = load_workbook('job_listings.xlsx')
            ws = wb.active

            # Need the index to add the 'date'
            last_line = ws.max_row

    except:
        raise Exception
    else:
        date_cell = ws.cell(row=last_line + 3, column=1)
        date_cell.font = styles.Font(name='Arial', italic=True, color='888888', size=10)
        # Fetch date
        today = date.today()
        formatted_date = today.strftime("%d/%m/%Y")
        date_cell.value = formatted_date

        ws.append(['-' * 100])
        # Add header
        ws.append(['Title', 'Company', 'Job', 'Salary', 'Location'])
        header_row = ws.max_row

        # format/style the heading
        for col in range(1, 6):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = styles.Font(name='Arial', bold=True, color='FFFFFF')
            cell.fill = styles.PatternFill('solid', start_color='4472C4')  # medium blue
            cell.alignment = styles.Alignment(horizontal='center')

        # Add the jobs
        for i, job in enumerate(all_jobs):
            ws.append(job)
            if i % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=ws.max_row, column=col).fill = styles.PatternFill('solid', start_color='DCE6F1')

        # Wrap the text in Job Column C
        first_data_row = header_row + 1
        for row in ws.iter_rows(min_row=first_data_row):
            row[2].alignment = styles.Alignment(wrap_text=True, vertical='top')

        # Set a fixed height row for readability
        for row_num in range(first_data_row, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 50

        # Add a border to the header row
        thin = Border(bottom=Side(style='medium'))
        for col in range(1, 6):
            ws.cell(row=header_row, column=col).border = thin

        wb.save('job_listings.xlsx')
